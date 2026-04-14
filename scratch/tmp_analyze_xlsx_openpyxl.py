import openpyxl, json
path = r'C:\Users\user_9nAJQ3l23\.qclaw\media\inbound\推广分析_23.03.2026---dc558f00-538d-47d2-b13c-227320c3c6f1.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
out = {'sheets': wb.sheetnames, 'preview': {}}
for s in wb.sheetnames:
    ws = wb[s]
    rows = list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 12)))
    cols = rows[0] if rows else []
    data = []
    for r in rows[1:11]:
        data.append({str(cols[i]): '' if (i >= len(r) or r[i] is None) else str(r[i]) for i in range(len(cols))})
    out['preview'][s] = {
        'shape': [ws.max_row, ws.max_column],
        'columns': [str(c) for c in cols],
        'head': data
    }
print(json.dumps(out, ensure_ascii=False))
