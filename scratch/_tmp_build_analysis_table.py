import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

order_path = r'C:\Users\user_9nAJQ3l23\.qclaw\media\inbound\店铺订单详情_3月---65789525-9bd2-4653-bd48-82a5c232fc25.xlsx'
promo_path = r'C:\Users\user_9nAJQ3l23\.qclaw\media\inbound\商品推广_分天数据_20260228至20260329---cd39f777-cfee-4128-9d99-1aa83cdd91b2.xlsx'
out_path = r'C:\Users\user_9nAJQ3l23\.qclaw\workspace\商品广告分析表_2026-03-30.xlsx'

odf = pd.read_excel(order_path)
odf['商品id'] = odf['商品id'].astype(str)
odf['支付时间'] = pd.to_datetime(odf['支付时间'], errors='coerce')
odf['用户实付金额(元)'] = pd.to_numeric(odf['用户实付金额(元)'], errors='coerce').fillna(0)
refund_kw = odf['售后状态'].astype(str).str.contains('退款', na=False)
ord_paid = odf[odf['支付时间'].notna() & ~odf['商品'].astype(str).str.contains('补收差价', na=False)].copy()
ord_prod = ord_paid.groupby(['商品id','商品']).agg(店铺订单数=('订单号','count'), 店铺销售额=('用户实付金额(元)','sum')).reset_index()
refund_prod = odf[refund_kw].groupby(['商品id','商品']).agg(退款订单数=('订单号','count'), 退款涉及金额=('用户实付金额(元)','sum')).reset_index()

pdf = pd.read_excel(promo_path, sheet_name=1)
pdf['日期_dt'] = pd.to_datetime(pdf['日期'], errors='coerce')
pdf = pdf[pdf['日期_dt'].notna()].copy()
pdf['商品ID'] = pdf['商品ID'].astype(str)
num_cols = ['净成交花费(元)','净交易额(元)','净成交笔数','曝光量','点击量','直接净交易额(元)','间接净交易额(元)','直接净成交笔数','间接净成交笔数']
for c in num_cols:
    pdf[c] = pd.to_numeric(pdf[c], errors='coerce').fillna(0)
prod = pdf.groupby(['商品ID','商品名称']).agg(
    广告花费=('净成交花费(元)','sum'),
    广告成交额=('净交易额(元)','sum'),
    广告订单数=('净成交笔数','sum'),
    直接成交额=('直接净交易额(元)','sum'),
    间接成交额=('间接净交易额(元)','sum'),
    直接订单数=('直接净成交笔数','sum'),
    间接订单数=('间接净成交笔数','sum'),
    曝光量=('曝光量','sum'),
    点击量=('点击量','sum')
).reset_index()
prod['ROAS'] = np.where(prod['广告花费']>0, prod['广告成交额']/prod['广告花费'], 0)
prod['CTR'] = np.where(prod['曝光量']>0, prod['点击量']/prod['曝光量']*100, 0)
prod['CPC'] = np.where(prod['点击量']>0, prod['广告花费']/prod['点击量'], 0)
prod = prod.merge(ord_prod, left_on='商品ID', right_on='商品id', how='left').drop(columns=['商品id'], errors='ignore')
prod = prod.merge(refund_prod, left_on='商品ID', right_on='商品id', how='left', suffixes=('','_refund')).drop(columns=['商品id'], errors='ignore')
prod = prod.fillna(0)
prod['广告成交占店铺销售比%'] = np.where(prod['店铺销售额']>0, prod['广告成交额']/prod['店铺销售额']*100, 0)


def judge(r):
    if r['广告花费'] > 50 and r['广告订单数'] == 0:
        return '有消耗无成交'
    if r['ROAS'] >= 8 and r['退款订单数'] <= 2:
        return '高效稳定'
    if r['ROAS'] >= 6 and r['退款订单数'] > 2:
        return '高效但退款偏高'
    if 4 <= r['ROAS'] < 6:
        return '中等效率'
    if 0 < r['ROAS'] < 4:
        return '低效转化弱'
    if r['广告花费'] == 0 and r['店铺销售额'] > 0:
        return '自然单为主'
    return '待观察'

def suggest(r):
    if r['广告花费'] > 50 and r['广告订单数'] == 0:
        return '暂停投放，检查主图/价格/详情页'
    if r['ROAS'] >= 8 and r['退款订单数'] <= 2:
        return '可加预算，继续放量'
    if r['ROAS'] >= 6 and r['退款订单数'] > 2:
        return '保留投放，优先处理退款问题'
    if 4 <= r['ROAS'] < 6:
        return '保留观察，优化素材和转化页'
    if 0 < r['ROAS'] < 4:
        return '降预算，重新测试定位'
    if r['广告花费'] == 0 and r['店铺销售额'] > 0:
        return '可视情况补投，先保自然转化'
    return '继续观察数据'

def action(r):
    if r['广告花费'] > 50 and r['广告订单数'] == 0:
        return '停投'
    if r['ROAS'] >= 8 and r['退款订单数'] <= 2:
        return '加预算'
    if r['ROAS'] >= 6:
        return '保留'
    if 4 <= r['ROAS'] < 6:
        return '观察'
    if 0 < r['ROAS'] < 4:
        return '降预算'
    if r['广告花费'] == 0 and r['店铺销售额'] > 0:
        return '自然单'
    return '观察'

prod['问题判断'] = prod.apply(judge, axis=1)
prod['优化建议'] = prod.apply(suggest, axis=1)
prod['操作建议'] = prod.apply(action, axis=1)
prod = prod.sort_values(['广告成交额','ROAS'], ascending=[False,False])

cols = ['商品ID','商品名称','广告花费','广告成交额','广告订单数','ROAS','直接成交额','间接成交额','曝光量','点击量','CTR','CPC','店铺销售额','店铺订单数','退款订单数','退款涉及金额','广告成交占店铺销售比%','问题判断','优化建议','操作建议']
prod[cols].to_excel(out_path, index=False, sheet_name='商品分析')

wb = load_workbook(out_path)
ws = wb['商品分析']
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, ws.max_column + 1):
    max_len = 0
    col_letter = get_column_letter(col)
    for cell in ws[col_letter]:
        v = '' if cell.value is None else str(cell.value)
        max_len = max(max_len, min(len(v), 40))
    ws.column_dimensions[col_letter].width = max_len + 2

money_cols = ['C','D','G','H','L','M','P','Q']
for col in money_cols:
    for cell in ws[col][1:]:
        cell.number_format = '0.00'
percent_cols = ['F','K','Q']
for col in percent_cols:
    for cell in ws[col][1:]:
        cell.number_format = '0.00'

fill_map = {
    '加预算':'C6E0B4',
    '保留':'FFF2CC',
    '观察':'FCE4D6',
    '降预算':'F4CCCC',
    '停投':'EA9999',
    '自然单':'D9EAD3'
}
for row in range(2, ws.max_row + 1):
    val = ws[f'T{row}'].value
    if val in fill_map:
        ws[f'T{row}'].fill = PatternFill('solid', fgColor=fill_map[val])

wb.save(out_path)
print(out_path)
